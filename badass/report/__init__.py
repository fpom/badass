import sys, tempfile, io, re

from functools import reduce
from operator import or_
from pathlib import Path
from collections import namedtuple, defaultdict
from zipfile import ZipFile, ZIP_STORED, ZIP_LZMA
from io import StringIO
from csv import DictReader, DictWriter
from werkzeug.utils import secure_filename

from .. import encoding
from ..db import connect

class Test (object) :
    _TEST = {"pass" : 0,
             "warn" : 1,
             "fail" : 2}
    _STAT = {v : k for k, v in _TEST.items()}
    NAMES = {}
    PASS = "PASS"
    WARN = "WARN"
    FAIL = "FAIL"
    def __init__ (self, val, status, num=None, txt="") :
        self.val = val
        self.status = status
        self.num = num
        self.txt = txt
        self.children = []
    def value (self) :
        if self.children :
            return sum(c.value() for c in self.children) / len(self.children)
        else :
            return self.val / float(max(self._STAT))
    def __iter__ (self) :
        if self.children :
            if self.num :
                yield self.num, self.val
            for child in self.children :
                yield from child
        else :
            yield self.num, self.val
    @classmethod
    def from_csv (cls, stream) :
        root = cls(0, "pass")
        nodes = {"" : root}
        names = {}
        for row in DictReader(stream) :
            num = row["test"]
            dot = f".{num}"
            nodes[dot] = cls(cls._TEST[row["status"]], row["status"], num, row["text"])
            nodes[dot.rsplit(".", 1)[0]].children.append(nodes[dot])
            if row["auto"] != "True" :
                names[num] = row["text"]
        if not cls.NAMES :
            cls.NAMES.update(names)
        else :
            for num in list(cls.NAMES) :
                if num not in names :
                    del cls.NAMES[num]
                else :
                    cls.NAMES[num] = cls._prefix(cls.NAMES[num], names[num])
        root.status = cls._STAT[max(c.val for c in root.children)]
        return root
    @classmethod
    def _prefix (cls, first, *rest) :
        cut = min(len(first), *(len(s) for s in rest))
        while not all(first[:cut] == s[:cut] for s in rest) :
            cut -= 1
        if cut < max(len(first), *(len(s) for s in rest)) :
            return first[:cut] + "…"
        else :
            return first[:cut]
    @classmethod
    def headers (cls) :
        yield from sorted(cls.NAMES.items())
    def __str__ (self) :
        out = io.StringIO()
        self._print(out)
        return out.getvalue()
    def _print (self, out, prefix=None, last=True) :
        status = getattr(self, self.status.upper(), "")
        label = f"{status} {self.txt}".rstrip()
        if prefix is None :
            out.write(f"{label}\n")
        elif last :
            out.write(f"{prefix} └─ {label}\n")
        else :
            out.write(f"{prefix} ├─ {label}\n")
        for child in self.children :
            if prefix is None :
                child._print(out, "", child is self.children[-1])
            elif last :
                child._print(out, prefix + "    ", child is self.children[-1])
            else :
                child._print(out, prefix + " │  ", child is self.children[-1])

class Report (object) :
    def __init__ (self, dbpath, groups, exercises) :
        # only load if necessary to speedup prog startup
        global Workbook, dataframe_to_rows, PatternFill, Alignment
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Alignment
        #
        DB, CFG, _, _ = connect(dbpath)
        exos_by_course = defaultdict(set)
        for ex in exercises :
            c, e = ex.split("/", 1)
            exos_by_course[c].add(e)
        self.xlsx_init()
        self.content = {}
        for c, exos in sorted(exos_by_course.items()) :
            for e in sorted(exos) :
                self.xlsx_new_ws(f"{c}-{e}")
                dbfilter = ((DB.users.id == DB.submissions.user)
                            & reduce(or_, (DB.users.group == g for g in groups))
                            & (DB.submissions.exercise == e)
                            & (DB.submissions.course == c))
                for row in DB(dbfilter).select(DB.users.firstname,
                                               DB.users.lastname,
                                               DB.users.studentid,
                                               DB.users.group,
                                               DB.submissions.date,
                                               DB.submissions.path) :
                    self.xlsx_add_row(row)
                    root = Path(row.submissions.path)
                    name = f"{row.users.lastname} {row.users.firstname}".title()
                    head = Path(secure_filename(name), *root.parts[1:])
                    self.content.update(self._walk(root, head, root))
                self.xlsx_done_ws()
    def _walk (self, root, head, sub) :
        if sub.is_dir() :
            for path in sub.iterdir() :
                if path.is_file() :
                    yield path, head / path.relative_to(root)
                elif path.is_dir() :
                    yield from self._walk(root, head, path)
    def save (self, path) :
        with ZipFile(path, "w", compression=ZIP_STORED) as zf :
            zf.writestr("report.xlsx", self.xlsx_data(),
                        compress_type=ZIP_LZMA, compresslevel=9)
            for cont, name in self.content.items() :
                if cont.suffix == ".zip" :
                    comp = {}
                else :
                    comp = {"compress_type" : ZIP_LZMA,
                            "compresslevel" : 9}
                zf.write(cont, name, **comp)
    def xlsx_init (self) :
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
    def xlsx_data (self) :
        with tempfile.NamedTemporaryFile(mode="w+b", suffix=".xlsx") as tmp :
            self.wb.save(tmp.name)
            tmp.seek(0)
            return tmp.read()
    _sheetname = re.compile("-*[^a-z0-9-]+-*", re.I)
    def xlsx_new_ws (self, name) :
        name = self._sheetname.sub("-", name)
        self.ws = self.wb.create_sheet(name)
        self.headers = ["student", "name", "group", "score", "best", "date", "report"]
        self.rows = []
        self.best = {}
    def xlsx_add_row (self, row) :
        name = f"{row.users.lastname.upper()} {row.users.firstname.title()}"
        path = Path(row.submissions.path)
        report = path / "report.zip"
        try :
            with ZipFile(report) as zf :
                test_data = io.StringIO(zf.read("report.csv").decode(**encoding))
        except :
            self.rows.append([row.users.studentid,
                              name,
                              row.users.group,
                              "CRASH",
                              False,
                              row.submissions.date])
            return
        test = Test.from_csv(test_data)
        try :
            permalink = (path / "permalink").read_text(**encoding)
        except :
            permalink = "missing"
        score = 1 - test.value()
        self.best[row.users.studentid] = max(score,
                                             self.best.get(row.users.studentid, 0))
        self.rows.append([row.users.studentid,
                          name,
                          row.users.group,
                          score,
                          False,
                          row.submissions.date,
                          permalink,
                          dict(test)])
    def _xlsx_add_row (self, row, values, styles=None, formats=None) :
        if not isinstance(styles, (list, tuple)) :
            styles = [styles] * len(values)
        if not isinstance(formats, (list, tuple)) :
            formats = [formats] * len(values)
        for col, (val, sty, fmt) in enumerate(zip(values, styles, formats), 1) :
            cell = self.ws.cell(column=col, row=row, value=val)
            if callable(sty) :
                sty(cell)
            elif sty is not None :
                cell.style = sty
            if callable(fmt) :
                fmt(cell)
            elif fmt is not None :
                cell.number_format = fmt
    def xlsx_done_ws (self) :
        self.rows.sort(key=lambda row: (row[0], row[5]))
        # write headers
        tests = list(Test.headers())
        self.headers.extend(f"{num}. {txt}" for num, txt in tests)
        self._xlsx_add_row(1, self.headers, styles=self._xlsx_style_head)
        # additional headers style
        self.ws.row_dimensions[1].height = 200
        self.ws.auto_filter.ref = self.ws.dimensions
        # write rows
        for num, row in enumerate(self.rows, 2) :
            if row[3] == "CRASH" :
                styles = "Bad"
                formats = [None, None, None, None,
                           self._xlsx_format_best,
                           self._xlsx_format_date]
            else :
                # update best column = row[4]
                # row[0] = studentid / row[3] = score
                best = row[4] = (row[3] == self.best.get(row[0], None))
                if best :
                    styles = ["Good" if hdr in self._HEAD_WIDTH
                              else self._xlsx_style_mark
                              for hdr in self.headers]
                else :
                    styles = [None if hdr in self._HEAD_WIDTH
                              else self._xlsx_style_mark
                              for hdr in self.headers]
                formats = [getattr(self, f"_xlsx_format_{hdr}", None)
                           for hdr in self.headers]
                test = row.pop(-1)
                row.extend(test[num] for num, _ in tests)
            self._xlsx_add_row(num, row, styles=styles, formats=formats)
    _HEAD_WIDTH = {"student" : 12,
                   "name" : 16,
                   "group" : 8,
                   "score" : 8,
                   "best" : 8,
                   "date" : 12,
                   "report" : 8}
    def _xlsx_style_head (self, cell) :
        cell.style = "Note"
        if cell.value not in self._HEAD_WIDTH :
            cell.alignment = Alignment(textRotation=90)
        width = self._HEAD_WIDTH.get(cell.value, 3)
        self.ws.column_dimensions[cell.column_letter].width = width
    def _xlsx_style_mark (self, cell) :
        if cell.value == 0 :
            cell.style = "Good"
        elif cell.value == 1 :
            cell.style = "Neutral"
        elif cell.value == 2 :
            cell.style = "Bad"
    def _xlsx_format_score (self, cell) :
        cell.number_format = "0%"
    def _xlsx_format_best (self, cell) :
        cell.number_format = "BOOLEAN"
    def _xlsx_format_date (self, cell) :
        cell.number_format = "MM-DD HH:MM"
    def _xlsx_format_report (self, cell) :
        if isinstance(cell.value, str) :
            if cell.value == "missing" :
                cell.style = "Bad"
            else :
                cell.hyperlink = cell.value
                cell.value = "link"
